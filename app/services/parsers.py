import csv
from typing import List, Dict, Any

try:
    # Optional dependency for reading .xlsx
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


# ==== Item Master Import (User-provided export format) ====
ITEM_MASTER_HEADERS = [
    "Division", "Item Code", "Description", "Unit", "Rate", "Region"
]

# Pivoted Item Master headers (export format)
PIVOT_REGION_COLUMNS = [
    "Dhaka Zone", "Mymensingh Zone", "Comilla Zone", "Sylhet Zone",
    "Khulna Zone", "Barisal Zone", "Gopalganj Zone", "Rajshahi Zone",
    "Rangpur Zone", "Chattogram Zone"
]

def parse_item_master_pivot_csv_text(text: str) -> List[Dict[str, Any]]:
    """Parse pivoted Item Master CSV text to a list of dicts expected by ItemParsed.

    Flexible header parsing:
    - Accept base headers using common synonyms and case-insensitive matching.
    - Treat any non-base columns (except optional SI. No and Organization) as region columns.
    """
    data: List[Dict[str, Any]] = []
    reader = csv.DictReader(text.splitlines())
    fieldnames_raw = reader.fieldnames or []
    # Normalize headers: strip and lower for matching, keep original for value access
    norm = lambda s: str(s or "").strip().lower()
    field_map = {norm(h): h for h in fieldnames_raw}

    # Synonyms for base columns
    base_synonyms = {
        "item_code": ["item code", "code", "itemcode", "item"],
        "division": ["major division", "Division", "division name"],
        "description": ["description", "item description", "desc"],
        "unit": ["unit", "units"],
        "organization": ["organization", "org","organisation"],
        "si_no": ["si. no", "si no", "serial", "sl", "sl."],
    }

    def resolve_header(key: str) -> str | None:
        for cand in base_synonyms[key]:
            if cand in field_map:
                return field_map[cand]
        return None

    item_code_h = resolve_header("item_code")
    division_h = resolve_header("division")
    description_h = resolve_header("description")
    unit_h = resolve_header("unit")
    org_h = resolve_header("organization")
    si_h = resolve_header("si_no")

    missing = [label for label, h in {
        "Item Code": item_code_h,
        "Major Division": division_h,
        "Description": description_h,
        "Unit": unit_h,
    }.items() if h is None]
    if missing:
        raise ValueError(f"Missing expected base headers in pivot Item Master CSV: {missing}")

    # Region headers = all headers minus base and optional SI/Organization
    base_set = {item_code_h, division_h, description_h, unit_h}
    optional_set = {si_h, org_h}
    region_headers = [h for h in fieldnames_raw if h not in base_set and h not in optional_set]
    # Backward compatibility: allow 'Cumilla Zone' as 'Comilla Zone'
    region_headers = ["Comilla Zone" if h == "Cumilla Zone" else h for h in region_headers]

    def clean_str(value) -> str:
        s = str(value).strip() if value is not None else ""
        return "" if s.lower() in ("none", "null", "-") else s

    def clean_unit(value) -> str | None:
        s = str(value).strip() if value is not None else ""
        return None if s == "" or s.lower() in ("none", "null", "-") else s

    for row in reader:
        division = clean_str(row.get(division_h))
        item_code = clean_str(row.get(item_code_h))
        description = clean_str(row.get(description_h))
        unit = clean_unit(row.get(unit_h))
        org_raw = row.get(org_h) if org_h else None
        organization = clean_str(org_raw) or "RHD"
        if not item_code and not description:
            continue

        for region in region_headers:
            rate_str = row.get(region)
            rate = None
            if rate_str not in (None, "", "-"):
                try:
                    rate = float(rate_str)
                except ValueError:
                    rate = None
            entry: Dict[str, Any] = {
                "division": division,
                "item_code": item_code,
                "item_description": description,
                "unit": unit,
                "rate": rate,
                "region": region,
                "organization": organization,
            }
            data.append(entry)
    return data

def parse_item_master_csv_text(text: str) -> List[Dict[str, Any]]:
    """Parse Item Master CSV text (string) to a list of dicts expected by ItemParsed."""
    data: List[Dict[str, Any]] = []
    reader = csv.DictReader(text.splitlines())
    # Basic header validation
    missing = [h for h in ITEM_MASTER_HEADERS if h not in reader.fieldnames]
    if missing:
        raise ValueError(f"Missing expected headers in Item Master CSV: {missing}")
    def clean_str(value) -> str:
        s = str(value).strip() if value is not None else ""
        return "" if s.lower() in ("none", "null", "-") else s

    def clean_unit(value) -> str | None:
        s = str(value).strip() if value is not None else ""
        return None if s == "" or s.lower() in ("none", "null", "-") else s

    for row in reader:
        rate_val = row.get("Rate")
        try:
            rate = float(rate_val) if rate_val not in (None, "", "-") else None
        except ValueError:
            rate = None
        entry = {
            "division": clean_str(row.get("Division")),
            "item_code": clean_str(row.get("Item Code")),
            "item_description": clean_str(row.get("Description")),
            "unit": clean_unit(row.get("Unit")),
            "rate": rate,
            "region": clean_str(row.get("Region")),
        }
        # Skip rows missing essential identifiers
        if not entry["item_code"] and not entry["item_description"]:
            continue
        data.append(entry)
    return data

def parse_item_master_xlsx_bytes(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Parse Item Master XLSX in memory to list of dicts expected by ItemParsed.
    
    Flexible header matching - accepts common synonyms and is case-insensitive.
    """
    if load_workbook is None:
        raise ImportError("openpyxl is not installed. Please add 'openpyxl' to requirements or install it.")
    from io import BytesIO
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.active
    
    # Get headers from first row
    headers_raw = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"DEBUG XLSX: Found {len(headers_raw)} headers: {headers_raw}")
    
    # Build case-insensitive header lookup
    norm = lambda s: str(s or "").strip().lower()
    norm_to_orig = {norm(h): h for h in headers_raw}
    
    # Define synonyms for each required column
    header_map = {
        "division": ["division", "major division", "div"],
        "item_code": ["item code", "code", "itemcode", "item", "item_code"],
        "item_description": ["description", "item description", "desc", "item desc"],
        "unit": ["unit", "units"],
        "rate": ["rate", "rate/unit"],
        "region": ["region", "zone", "area"],
    }
    
    # Find which column index corresponds to each field
    field_to_idx = {}
    for field_name, synonyms in header_map.items():
        for synonym in synonyms:
            if synonym in norm_to_orig:
                orig_header = norm_to_orig[synonym]
                field_to_idx[field_name] = headers_raw.index(orig_header)
                print(f"DEBUG XLSX: Matched '{field_name}' to column '{orig_header}' (index {field_to_idx[field_name]})")
                break
    
    # Check we found all required fields
    required = ["division", "item_code", "item_description", "unit", "rate", "region"]
    missing = [f for f in required if f not in field_to_idx]
    if missing:
        raise ValueError(f"Missing required columns: {missing}. Found: {headers_raw}")
    
    data: List[Dict[str, Any]] = []
    
    def clean_str(value) -> str:
        s = str(value).strip() if value else ""
        return "" if s.lower() in ("none", "null", "-") else s
    
    def clean_unit(value) -> str | None:
        s = str(value).strip() if value else ""
        return None if s == "" or s.lower() in ("none", "null", "-") else s
    
    # Parse data rows
    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            # Extract values using the column indexes we found
            div_val = row[field_to_idx["division"]].value if field_to_idx["division"] < len(row) else None
            code_val = row[field_to_idx["item_code"]].value if field_to_idx["item_code"] < len(row) else None
            desc_val = row[field_to_idx["item_description"]].value if field_to_idx["item_description"] < len(row) else None
            unit_val = row[field_to_idx["unit"]].value if field_to_idx["unit"] < len(row) else None
            rate_val = row[field_to_idx["rate"]].value if field_to_idx["rate"] < len(row) else None
            region_val = row[field_to_idx["region"]].value if field_to_idx["region"] < len(row) else None
            
            # Parse rate
            try:
                rate = float(rate_val) if rate_val not in (None, "", "-") else None
            except (ValueError, TypeError):
                rate = None
            
            entry = {
                "division": clean_str(div_val),
                "item_code": clean_str(code_val),
                "item_description": clean_str(desc_val),
                "unit": clean_unit(unit_val),
                "rate": rate,
                "region": clean_str(region_val),
            }
            
            # Skip empty rows
            if not entry["item_code"] and not entry["item_description"]:
                continue
            
            data.append(entry)
        except Exception as e:
            # Log but continue on row parsing errors
            print(f"DEBUG XLSX: Error parsing row {row_num}: {e}")
            continue
    
    print(f"DEBUG XLSX: Successfully parsed {len(data)} rows from XLSX")
    return data

def parse_item_master_pivot_xlsx_bytes(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Parse pivoted Item Master XLSX in memory to list of dicts expected by ItemParsed.

    Flexible header parsing with case-insensitive synonyms and dynamic region columns.
    """
    if load_workbook is None:
        raise ImportError("openpyxl is not installed. Please add 'openpyxl' to requirements or install it.")
    from io import BytesIO
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.active
    headers = [cell.value if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_index = {str(h).strip(): idx for idx, h in enumerate(headers)}
    # Build normalized lookup for synonyms
    norm = lambda s: str(s or "").strip().lower()
    norm_index = {norm(h): idx for h, idx in header_index.items()}
    base_synonyms = {
        "item_code": ["item code", "code", "itemcode", "item"],
        "division": ["major division", "division", "division name"],
        "description": ["description", "item description", "desc"],
        "unit": ["unit", "units"],
        "organization": ["organization", "org"],
        "si_no": ["si. no", "si no", "serial", "sl", "sl."],
    }
    def find_idx(key: str) -> int | None:
        for cand in base_synonyms[key]:
            if cand in norm_index:
                return norm_index[cand]
        return None
    idx_item_code = find_idx("item_code")
    idx_division = find_idx("division")
    idx_description = find_idx("description")
    idx_unit = find_idx("unit")
    idx_org = find_idx("organization")
    idx_si = find_idx("si_no")
    missing_base = [label for label, idx in {
        "Item Code": idx_item_code,
        "Major Division": idx_division,
        "Description": idx_description,
        "Unit": idx_unit,
    }.items() if idx is None]
    if missing_base:
        # Check if we are in a simple linear format (not pivoted) that failed detection
        # But this function is specifically for pivoted... 
        # Let's try to be more flexible? No, if basic columns are missing, we can't map data.
        raise ValueError(f"Missing expected base headers in pivot Item Master XLSX: {missing_base}. Found: {headers}")

    # Region headers: any headers not part of base or optional SI/Organization
    # We also need to handle comma-separated regions in headers (e.g. "Dhaka, Mymensingh")
    # This means one column provides rate for MULTIPLE regions.
    
    # Map column index -> List[region_name]
    col_idx_to_regions: Dict[int, List[str]] = {}
    
    known_base_indices = {idx_item_code, idx_division, idx_description, idx_unit, idx_org, idx_si}
    
    for idx, h in enumerate(headers):
        if idx in known_base_indices:
            continue
            
        h_str = str(h).strip()
        if not h_str: 
            continue
            
        # Split by comma or slash to handle "Dhaka, Mymensingh" or "Dhaka/Mymensingh"
        # Also handle "Zone" suffix if present
        parts = [p.strip() for p in h_str.replace("/", ",").split(",")]
        
        regions_for_col = []
        for p in parts:
            if not p: continue
            # Normalize region name: add " Zone" if missing and it looks like a city name
            # But let's trust the input mostly, just normalize casing/spaces
            p_clean = p
            # Common RHD zones: Dhaka, Mymensingh, Comilla, Sylhet, Chittagong, Khulna, Barisal, Rajshahi, Rangpur, Gopalganj
            # If header is just "Dhaka", map to "Dhaka Zone" if that's the standard in DB?
            # DB uses "Default" or whatever user imports. Let's append " Zone" if not present to match RHD standard if needed?
            # Actually, let's keep it as is, but maybe map "Cumilla" -> "Comilla"
            if "Cumilla" in p_clean:
                p_clean = p_clean.replace("Cumilla", "Comilla")
            
            # If the header doesn't end with "Zone", should we add it? 
            # The previous code assumed "Dhaka Zone". 
            # If the user file has "Dhaka, Mymensingh", they probably mean "Dhaka Zone" and "Mymensingh Zone"
            # Let's heuristically add " Zone" if it's likely a region name and doesn't have it
            if "Zone" not in p_clean and p_clean in ["Dhaka", "Mymensingh", "Comilla", "Sylhet", "Chittagong", "Chattogram", "Khulna", "Barisal", "Barishal", "Rajshahi", "Rangpur", "Gopalganj"]:
                 p_clean += " Zone"
            
            regions_for_col.append(p_clean)
            
        if regions_for_col:
            col_idx_to_regions[idx] = regions_for_col

    data: List[Dict[str, Any]] = []
    def clean_str(value) -> str:
        s = str(value).strip() if value is not None else ""
        return "" if s.lower() in ("none", "null", "-") else s

    def clean_unit(value) -> str | None:
        s = str(value).strip() if value is not None else ""
        return None if s == "" or s.lower() in ("none", "null", "-") else s

    for row in ws.iter_rows(min_row=2):
        # Extract base values
        division = clean_str(row[idx_division].value if idx_division is not None else None)
        item_code = clean_str(row[idx_item_code].value if idx_item_code is not None else None)
        description = clean_str(row[idx_description].value if idx_description is not None else None)
        unit_cell = row[idx_unit] if idx_unit is not None else None
        unit = clean_unit(unit_cell.value if unit_cell else None)
        org_cell = row[idx_org] if idx_org is not None else None
        organization = clean_str(org_cell.value if org_cell else None)

        if not item_code and not description:
            continue

        # Iterate over region columns
        for col_idx, region_list in col_idx_to_regions.items():
            if col_idx >= len(row): continue
            
            rate_cell = row[col_idx]
            rate_val = rate_cell.value if rate_cell else None
            
            rate = None
            if rate_val not in (None, "-", ""):
                try:
                    rate = float(rate_val)
                except (ValueError, TypeError):
                    rate = None
            
            # If rate is present (or even if None, to create entry?), create entry for EACH region in this column
            # Usually if rate is None/Empty, we might skip or import as None.
            # Let's import even if None so we know the item exists in that region? 
            # Or better, if rate is None, maybe the item doesn't exist there?
            # Standard practice: Import if valid row. Rate can be 0 or None.
            
            for region_name in region_list:
                entry = {
                    "division": division,
                    "item_code": item_code,
                    "item_description": description,
                    "unit": unit,
                    "rate": rate,
                    "region": region_name,
                    "organization": organization,
                }
                data.append(entry)
                
    return data

# Removed CLI demo block to avoid unused runtime code in production
