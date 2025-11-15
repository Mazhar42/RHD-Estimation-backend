import csv
import json
import os
from typing import List, Dict, Any

try:
    # Optional dependency for reading .xlsx
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

def parse_rates_csv(file_path):
    """
    Reads the Rates.csv file, parses it, and transforms it into a list of dictionaries.
    Each dictionary represents an item's rate in a specific region.
    """
    data = []
    
    # Define the region columns and their corresponding keys in the output
    region_columns = [
        "Dhaka Zone", "Mymensingh Zone", "Cumilla Zone", "Sylhet Zone",
        "Khulna Zone", "Barisal Zone", "Gopalganj Zone", "Rajshahi Zone",
        "Rangpur Zone", "Chattogram Zone"
    ]

    with open(file_path, mode='r', encoding='utf-8') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        for row in csv_reader:
            # Extract common fields
            common_data = {
                "division": row["Major Division"],
                "item_code": row["Item Code"],
                "item_description": row["Description"],
                "unit": row["Unit"]
            }

            # Iterate through region-specific rates
            for region in region_columns:
                rate_str = row[region]
                rate = None
                if rate_str and rate_str != '-':
                    try:
                        rate = float(rate_str)
                    except ValueError:
                        # Handle cases where rate might not be a valid float
                        rate = None # Or 0.0, depending on desired behavior
                
                entry = common_data.copy()
                entry["region"] = region
                entry["rate"] = rate
                data.append(entry)
    return data

def parse_rates_xlsx(file_path: str) -> List[Dict[str, Any]]:
    """
    Reads the RatesExcel.xlsx file and transforms it into a list of dictionaries,
    matching the schema used by parse_rates_csv.

    Requires openpyxl to be installed.
    """
    if load_workbook is None:
        raise ImportError("openpyxl is not installed. Please add 'openpyxl' to requirements or install it.")

    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb.active

    # Extract headers from the first row
    headers = [cell.value if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Map header names to column indices
    header_index = {str(h).strip(): idx for idx, h in enumerate(headers)}

    required_headers = [
        "SI. No", "Item Code", "Major Division", "Description", "Unit",
        "Dhaka Zone", "Mymensingh Zone", "Cumilla Zone", "Sylhet Zone",
        "Khulna Zone", "Barisal Zone", "Gopalganj Zone", "Rajshahi Zone",
        "Rangpur Zone", "Chattogram Zone",
    ]

    # Validate headers exist
    missing = [h for h in required_headers if h not in header_index]
    if missing:
        raise ValueError(f"Missing expected headers in XLSX: {missing}")

    region_columns = [
        "Dhaka Zone", "Mymensingh Zone", "Cumilla Zone", "Sylhet Zone",
        "Khulna Zone", "Barisal Zone", "Gopalganj Zone", "Rajshahi Zone",
        "Rangpur Zone", "Chattogram Zone",
    ]

    data: List[Dict[str, Any]] = []

    # Iterate over data rows
    for row in ws.iter_rows(min_row=2):
        # Read common fields safely
        def val(hname):
            idx = header_index[hname]
            cell = row[idx]
            return cell.value if cell.value is not None else None

        common_data = {
            "division": str(val("Major Division") or "").strip(),
            "item_code": str(val("Item Code") or "").strip(),
            "item_description": str(val("Description") or "").strip(),
            "unit": str(val("Unit") or "").strip() or None,
        }

        # Skip empty lines (no item code or description)
        if not common_data["item_code"] and not common_data["item_description"]:
            continue

        for region in region_columns:
            rate_cell = val(region)
            rate = None
            if rate_cell not in (None, "-"):
                try:
                    rate = float(rate_cell)
                except (ValueError, TypeError):
                    rate = None
            entry = common_data.copy()
            entry["region"] = region
            entry["rate"] = rate
            data.append(entry)

    return data

def parse_rates(file_path: str) -> List[Dict[str, Any]]:
    """
    Dispatch parser based on file extension.
    Supports .csv via csv.DictReader and .xlsx via openpyxl.
    """
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return parse_rates_csv(file_path)
    elif ext in (".xlsx", ".xlsm"):
        return parse_rates_xlsx(file_path)
    else:
        raise ValueError(f"Unsupported file extension: {ext}")


# ==== Item Master Import (User-provided export format) ====
ITEM_MASTER_HEADERS = [
    "Division", "Item Code", "Description", "Unit", "Rate", "Region"
]

def parse_item_master_csv_text(text: str) -> List[Dict[str, Any]]:
    """Parse Item Master CSV text (string) to a list of dicts expected by ItemParsed."""
    data: List[Dict[str, Any]] = []
    reader = csv.DictReader(text.splitlines())
    # Basic header validation
    missing = [h for h in ITEM_MASTER_HEADERS if h not in reader.fieldnames]
    if missing:
        raise ValueError(f"Missing expected headers in Item Master CSV: {missing}")
    for row in reader:
        rate_val = row.get("Rate")
        try:
            rate = float(rate_val) if rate_val not in (None, "", "-") else None
        except ValueError:
            rate = None
        entry = {
            "division": (row.get("Division") or "").strip(),
            "item_code": (row.get("Item Code") or "").strip(),
            "item_description": (row.get("Description") or "").strip(),
            "unit": (row.get("Unit") or "").strip() or None,
            "rate": rate,
            "region": (row.get("Region") or "").strip(),
        }
        # Skip rows missing essential identifiers
        if not entry["item_code"] and not entry["item_description"]:
            continue
        data.append(entry)
    return data

def parse_item_master_xlsx_bytes(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Parse Item Master XLSX in memory to list of dicts expected by ItemParsed."""
    if load_workbook is None:
        raise ImportError("openpyxl is not installed. Please add 'openpyxl' to requirements or install it.")
    from io import BytesIO
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.active
    # Headers
    headers = [cell.value if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_index = {str(h).strip(): idx for idx, h in enumerate(headers)}
    missing = [h for h in ITEM_MASTER_HEADERS if h not in header_index]
    if missing:
        raise ValueError(f"Missing expected headers in Item Master XLSX: {missing}")
    data: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2):
        def val(hname):
            idx = header_index[hname]
            cell = row[idx]
            return cell.value if cell.value is not None else None
        rate_cell = val("Rate")
        try:
            rate = float(rate_cell) if rate_cell not in (None, "", "-") else None
        except (ValueError, TypeError):
            rate = None
        entry = {
            "division": str(val("Division") or "").strip(),
            "item_code": str(val("Item Code") or "").strip(),
            "item_description": str(val("Description") or "").strip(),
            "unit": (str(val("Unit")) if val("Unit") is not None else "").strip() or None,
            "rate": rate,
            "region": str(val("Region") or "").strip(),
        }
        if not entry["item_code"] and not entry["item_description"]:
            continue
        data.append(entry)
    return data

if __name__ == "__main__":
    # Assuming the script is run from the project root or backend directory
    # Adjust the path to Rates.csv accordingly
    current_dir = os.path.dirname(__file__)
    csv_file_path = os.path.join(current_dir, "..", "Rates.csv")
    xlsx_file_path = os.path.join(current_dir, "..", "RatesExcel.xlsx")
    
    # Ensure the path is correct
    if not os.path.exists(csv_file_path):
        print(f"Error: Rates.csv not found at {csv_file_path}")
    else:
        # Prefer CSV if available; otherwise try XLSX
        file_to_parse = csv_file_path if os.path.exists(csv_file_path) else xlsx_file_path
        parsed_data = parse_rates(file_to_parse)
        print(f"Parsed file: {file_to_parse}")
        # For demonstration, print the first few entries and total count
        print(f"Total entries parsed: {len(parsed_data)}")
        print(json.dumps(parsed_data[:10], indent=2)) # Print first 10 entries
